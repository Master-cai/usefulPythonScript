# -*-coding:utf-8-*-
import os
import openpyxl




class XlsxKiller():

    def __init__(self, fileList, workPath, nameColNum, nameList, output, oldRowNum):
        self._fileList = fileList
        self._workPath = workPath
        self._nameColNum = nameColNum
        self._nameList = nameList
        self._output = output
        self._oldRowNum = oldRowNum

    

    def get_row_data(self):  # get the row data needed
        data_list = []
        with open(self._nameList) as f:
            nameList = f.readline().split('，')

        for file in self._fileList:
            data = openpyxl.load_workbook(os.path.join(self._workPath, file))
            table = data.worksheets[0]
            for i in range(3, table.max_row): # list(table.rows)[3][2].value
                name = list(table.rows)[i][2].value  # get name from row data
                if name != None and name in nameList:
                    data_list.append(list(table.rows)[i])
        # print(data_list)
        return data_list


    def check(self, nameList, name):  # used to check the name in the name list
        if name in nameList:
            nameList.remove(name)
            print('file {} has been merged'.format(name))

    def write_rows(self, rowsData):
        
        workBook = openpyxl.load_workbook(filename=self._output)
        workSheet = workBook.worksheets[0]
        with open(self._nameList) as f:
            nameList = f.readline().split('，')
        i = self._oldRowNum
        for index, row in enumerate(workSheet.rows):
            if index >= self._oldRowNum and rowsData:
                for j in range(len(row)):
                    row[j].value = rowsData[0][j].value
                self.check(nameList, rowsData[0][2].value)
                del rowsData[0]
        workBook.save('output.xlsx')
        return nameList


