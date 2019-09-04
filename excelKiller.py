# -*-coding:utf-8-*-
import os

import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
import argparse


def getlist(workSpaceDir):  # get the excel file list
    fileNames = os.listdir(workSpaceDir)
    seachedFiles = []
    for fileName in fileNames:
        if fileName.endswith('.xls'):
            seachedFiles.append(fileName)
            fileType = '.xls'
        if fileName.endswith('.xlsx'):
            seachedFiles.append(fileName)
            fileType = '.xlsx'
    return seachedFiles, fileType


def get_row_data(fileList, workPath, nameColNum, nameList):  # get the row data needed
    '''
    :param fileList: the file that the data exist
    :param workPath: the dictionary the excel files exist.
    :param nameColNum: the column number of the name column
    :return data_list: return a list that include all the data in rows.
    '''
    data_list = []
    for file in fileList:
        # name_true = file.split('.')[0]  # get name from file name
        data = xlrd.open_workbook(os.path.join(workPath, file))
        table = data.sheets()[0]
        for i in range(table.nrows):
            name = table.row(i)[nameColNum].value  # get name from row data
            if name in nameList:
                data_list.append(table.row(i))

    return data_list


def write_rows(rowsData, outPut, nameColNum, oldRowNum, nameList):
    '''
    :param rowsData: the row data to be wrote
    :param outPut: the file you want
    :param nameColNum: the column number that the name exist
    :param oldRowNum: the row number you want to begin
    :param nameList: all the name you want to check should be saved in the file.
    :return: return a list included all the name that didn`t found.
    '''
    workBook = copy(xlrd.open_workbook(outPut))
    workSheet = workBook.get_sheet(0)

    # set borders styles
    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    borders.bottom_colour = 0x3A
    style = xlwt.XFStyle()
    style.borders = borders

    i = oldRowNum
    with open(nameList) as f:
        nameList = f.readline().split('，')

    for row in rowsData:
        for j in range(len(row)):
            workSheet.write(i, j, str(row[j].value), style)
        check(nameList, row[nameColNum].value)
        i += 1
    workBook.save(outPut)
    return nameList


def addarg():
    parser = argparse.ArgumentParser()
    parser.add_argument('-w', '--workPath', type=str,
                        help='the path of your excels', default=os.getcwd())
    parser.add_argument('-o', '--oldRowNum', type=int,
                        help='the row number you want to begin', default=2)
    parser.add_argument('-c', '--column', type=int,
                        help='the column number of the name', default=3)
    parser.add_argument('-O', '--outPutFiles', type=str, help='the file you want ro save the output, it must exist',
                        default='output.xls')
    args = parser.parse_args()
    return args


def check(nameList, name):  # used to check the name in the name list
    if name in nameList:
        nameList.remove(name)
        print('file {} has been merged'.format(name))


def xls_killer(fileList, path, nameColNum, nameList, output, oldRowNum):
    rows = get_row_data(fileList, path, nameColNum, nameList)
    unDoneList = write_rows(rows, output, nameColNum, oldRowNum, nameList)
    return unDoneList

def xlsx_killer(fileList, path, nameColNum, nameList, output, oldRowNum):
    data_list = []
    for file in fileList:
        # name_true = file.split('.')[0]  # get name from file name
        data = openpyxl.load_workbook(os.path.join(path, file))
        table = data.worksheets[0]
        for i in range(table.max_row):
            # name = table.row(i)[nameColNum].value  # get name from row data
            print(list(table.rows)[i])
            
            name = table.cell(row=4, column=nameColNum+1).value
            print(os.path.join(path, file))
            print(name)
            if name in nameList:
                # data_list.append(table.row(i))
                data_list.append(list(table.rows)[i])
# end get_row_data

    workBook = openpyxl.load_workbook(output)
    workSheet = workBook.worksheets[0]
    i = oldRowNum
    with open(nameList) as f:
        nameList = f.readline().split('，')

    for row in data_list:
        workSheet.insert_rows(oldRowNum)
        check(nameList, row[nameColNum].value)
    #     i += 1
    workBook.save(filename=output)
    return nameList

    


def main():
    args = addarg()
    output = args.outPutFiles.strip('\'')
    path = args.workPath.strip('\'')
    nameList = 'nameList.txt'
    nameColNum = args.column
    oldRowNum = args.oldRowNum
    fileList, fileType = getlist(path)
    if fileType == '.xls':
        unDoneList = xls_killer(fileList, path, nameColNum, nameList, output, oldRowNum)
    if fileType == '.xlsx':
        unDoneList = xlsx_killer(fileList, path, nameColNum, nameList, output, oldRowNum)

    print('*'*20 + 'unDone' + '*'*20)
    for ele in unDoneList:
        print(ele+' unDone')


if __name__ == '__main__':
    main()
